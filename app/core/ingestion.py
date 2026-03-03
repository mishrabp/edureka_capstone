"""Document ingestion: load, parse, and chunk PDF/TXT/CSV/Excel files."""
from __future__ import annotations

import logging
import uuid
from pathlib import Path
from typing import List

import openpyxl
from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_community.document_loaders import (
    PyPDFLoader,
    TextLoader,
    CSVLoader,
)

from app.config import get_settings

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".pdf", ".txt", ".csv", ".xlsx", ".xls"}


def _load_excel(file_path: str, filename: str) -> List[Document]:
    """
    Load an Excel workbook using openpyxl and convert every sheet into
    a single Document whose content is a plain-text table representation.
    """
    try:
        # data_only=True evaluates formulas; sometimes triggers msoffcrypto check
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as exc:
        logger.warning("Failed to load Excel with data_only=True (%s); trying raw...", exc)
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
        except Exception as final_exc:
            logger.error("Excel load failed: %s", final_exc)
            raise

    docs: List[Document] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Build a readable text block: header + data rows
            lines: List[str] = [f"Sheet: {sheet_name}"]
            header_row = rows[0]
            header = [str(c) if c is not None else "" for c in header_row]
            lines.append(" | ".join(header))
            lines.append("-" * len(" | ".join(header)))

            for row in rows[1:]:
                cells = [str(c) if c is not None else "" for c in row]
                # Skip completely empty rows
                if any(c.strip() for c in cells):
                    lines.append(" | ".join(cells))

            content = "\n".join(lines)
            docs.append(
                Document(
                    page_content=content,
                    metadata={"sheet": sheet_name, "source": filename},
                )
            )
    finally:
        wb.close()

    return docs


def _get_loader(file_path: str, filename: str):
    """Return the appropriate loader for a given file type."""
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        return PyPDFLoader(file_path)
    elif ext == ".txt":
        return TextLoader(file_path, encoding="utf-8")
    elif ext == ".csv":
        return CSVLoader(file_path, encoding="utf-8")
    else:
        raise ValueError(f"Unsupported file type: {ext}. Supported: {SUPPORTED_EXTENSIONS}")


def load_and_chunk(file_path: str, filename: str, doc_id: str) -> List[Document]:
    """
    Load a document from disk and split it into chunks for embedding.

    Args:
        file_path: Absolute path to the file on disk.
        filename:  Original filename (used for metadata and extension detection).
        doc_id:    Unique identifier for this document.

    Returns:
        List of LangChain Document objects with enriched metadata.
    """
    settings = get_settings()
    ext = Path(filename).suffix.lower()

    logger.info("Loading document: %s (type=%s)", filename, ext)

    try:
        if ext in {".xlsx", ".xls"}:
            raw_docs = _load_excel(file_path, filename)
        else:
            loader = _get_loader(file_path, filename)
            raw_docs = loader.load()
    except Exception as exc:
        logger.error("Failed to load %s: %s", filename, exc)
        raise

    if not raw_docs:
        raise ValueError(f"No content extracted from {filename}.")

    # Attach consistent metadata to every raw page/element
    for doc in raw_docs:
        doc.metadata.setdefault("source", filename)
        doc.metadata["filename"] = filename
        doc.metadata["doc_id"] = doc_id
        doc.metadata["file_type"] = ext.lstrip(".")

    # Split into chunks
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=settings.chunk_size,
        chunk_overlap=settings.chunk_overlap,
        separators=["\n\n", "\n", ". ", " ", ""],
    )
    chunks = splitter.split_documents(raw_docs)

    # Assign a unique chunk ID to each chunk
    for i, chunk in enumerate(chunks):
        chunk.metadata["chunk_id"] = f"{doc_id}_{i}"

    logger.info("Chunked %s into %d chunks", filename, len(chunks))
    return chunks


def generate_doc_id() -> str:
    """Generate a unique document ID."""
    return str(uuid.uuid4())
