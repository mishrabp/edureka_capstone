"""
generate_samples.py
Generates realistic healthcare sample files for testing the Agentic RAG app.
  - sample_clinical_guidelines.pdf   (clinical protocol document)
  - sample_drug_interactions.csv     (drug interaction database)
  - sample_patient_records.xlsx      (patient records + prescriptions)
  - sample_treatment_notes.txt       (doctor's clinical notes)
"""
import os
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import csv

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "sample_data")
os.makedirs(OUTPUT_DIR, exist_ok=True)


# ─────────────────────────────────────────────────────────────────────────────
# 1. PDF — Clinical Guidelines
# ─────────────────────────────────────────────────────────────────────────────
def make_pdf():
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    def heading(text, size=14):
        pdf.set_font("Helvetica", "B", size)
        pdf.set_fill_color(220, 235, 245)
        pdf.cell(0, 10, text, ln=True, fill=True)
        pdf.ln(2)

    def body(text, size=10):
        pdf.set_font("Helvetica", "", size)
        pdf.multi_cell(0, 6, text)
        pdf.ln(2)

    # Cover
    pdf.set_font("Helvetica", "B", 20)
    pdf.cell(0, 15, "City General Hospital", ln=True, align="C")
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Clinical Treatment Guidelines — 2025 Edition", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 8, "Department of Internal Medicine | Version 4.2", ln=True, align="C")
    pdf.ln(10)

    heading("1. Hypertension Management Protocol")
    body(
        "Hypertension is defined as systolic blood pressure >= 130 mmHg or diastolic BP >= 80 mmHg "
        "(ACC/AHA 2025 Guidelines). First-line treatment includes lifestyle modification for all stages.\n\n"
        "Stage 1 (130-139 / 80-89 mmHg): Lifestyle changes for 3-6 months. If no improvement, initiate "
        "antihypertensive monotherapy. Preferred agents: ACE inhibitors (e.g., Lisinopril 10mg OD), "
        "ARBs (e.g., Losartan 50mg OD), or Thiazide diuretics (e.g., Hydrochlorothiazide 12.5mg OD).\n\n"
        "Stage 2 (>= 140/90 mmHg): Dual-agent therapy recommended. Combine ACE inhibitor + Calcium "
        "Channel Blocker (e.g., Amlodipine 5mg OD). Monitor BP every 4 weeks until controlled.\n\n"
        "Contraindications: ACE inhibitors are contraindicated in pregnancy (Category D). ARBs should "
        "not be used concurrently with ACE inhibitors due to increased risk of hyperkalaemia and AKI."
    )

    heading("2. Type 2 Diabetes Mellitus Protocol")
    body(
        "Diagnosis: Fasting blood glucose >= 126 mg/dL, HbA1c >= 6.5%, or 2-hour OGTT >= 200 mg/dL.\n\n"
        "First-Line Therapy: Metformin 500mg BD with meals (titrate to 2000mg/day). Contraindicated in "
        "eGFR < 30 mL/min/1.73m2 and active liver disease.\n\n"
        "Second-Line Options:\n"
        "  - SGLT2 inhibitors (Empagliflozin 10mg OD) — preferred in patients with HF or CKD.\n"
        "  - GLP-1 agonists (Semaglutide 0.5mg SC weekly) — preferred for weight reduction.\n"
        "  - DPP-4 inhibitors (Sitagliptin 100mg OD) — weight neutral, good tolerability.\n\n"
        "Monitoring: HbA1c every 3 months until target (< 7.0%), then every 6 months. "
        "Annual renal function (eGFR, urine ACR), lipid panel, and retinal screening."
    )

    heading("3. Antibiotic Prescribing Guidelines — Respiratory Infections")
    body(
        "Community-Acquired Pneumonia (CAP):\n"
        "  Outpatient (mild): Amoxicillin 1g TDS x 5 days OR Doxycycline 100mg BD x 5 days.\n"
        "  Inpatient (moderate): Amoxicillin-Clavulanate 1.2g IV TDS + Clarithromycin 500mg BD.\n"
        "  ICU (severe): Piperacillin-Tazobactam 4.5g IV TDS + Azithromycin 500mg IV OD.\n\n"
        "Penicillin Allergy: Replace Amoxicillin with Levofloxacin 750mg OD (5 days, outpatient) "
        "or Meropenem 1g IV TDS (severe inpatient).\n\n"
        "Duration: Minimum 5 days. Extend to 7 days if Legionella or Staphylococcal pneumonia.\n\n"
        "Note: Blood cultures MUST be drawn before initiating antibiotics in all inpatients."
    )

    heading("4. Sepsis — Early Recognition and Management (Sepsis-3)")
    body(
        "Definition: Life-threatening organ dysfunction caused by dysregulated host response to infection.\n"
        "SOFA score increase >= 2 points from baseline in the presence of suspected infection.\n\n"
        "Hour-1 Bundle (complete within 60 minutes of recognition):\n"
        "  1. Measure lactate. Re-measure if initial lactate > 2 mmol/L.\n"
        "  2. Obtain blood cultures (2 sets) BEFORE antibiotics.\n"
        "  3. Administer broad-spectrum antibiotics (e.g., Meropenem 1g IV + Vancomycin 25mg/kg).\n"
        "  4. Begin rapid infusion of 30 mL/kg IV crystalloid (Normal Saline or Lactated Ringer's).\n"
        "  5. Apply vasopressors if MAP < 65 mmHg despite resuscitation. Norepinephrine is first-line.\n\n"
        "Septic Shock: Refractory hypotension requiring vasopressors. Add Hydrocortisone 200mg/day "
        "IV if norepinephrine dose > 0.25 mcg/kg/min."
    )

    heading("5. Drug Contraindications in Renal Failure (eGFR < 30)")
    body(
        "The following medications require dose adjustment or are contraindicated in severe renal impairment:\n\n"
        "AVOID entirely:\n"
        "  - Metformin (risk of lactic acidosis)\n"
        "  - NSAIDs including Ibuprofen, Naproxen (further deterioration of renal function)\n"
        "  - Nitrofurantoin (ineffective and accumulates)\n"
        "  - Direct oral anticoagulants: Dabigatran (renally cleared — use Warfarin instead)\n\n"
        "Dose-reduce:\n"
        "  - Enoxaparin: reduce to 1 mg/kg SC OD (from normal BD)\n"
        "  - Gabapentin: max 300mg OD (normal dose up to 3600mg/day)\n"
        "  - Ciprofloxacin: 250-500mg BD (reduce by 50%)\n\n"
        "Monitor closely:\n"
        "  - Digoxin, Lithium — narrow therapeutic index; check levels weekly."
    )

    out = os.path.join(OUTPUT_DIR, "sample_clinical_guidelines.pdf")
    pdf.output(out)
    print(f"  ✓ PDF created: {out}")


# ─────────────────────────────────────────────────────────────────────────────
# 2. CSV — Drug Interaction Database
# ─────────────────────────────────────────────────────────────────────────────
def make_csv():
    rows = [
        ["Drug A", "Drug B", "Severity", "Interaction Effect", "Clinical Action"],
        ["Warfarin", "Aspirin", "Major", "Increased bleeding risk — additive anticoagulant effect", "Avoid combination; if necessary, use lowest aspirin dose and monitor INR closely"],
        ["Metformin", "Contrast Dye (IV)", "Major", "Risk of contrast-induced nephropathy and lactic acidosis", "Hold Metformin 48h before and after IV contrast; check renal function before restarting"],
        ["Lisinopril", "Potassium supplements", "Moderate", "Hyperkalaemia — ACE inhibitor reduces potassium excretion", "Monitor serum potassium every 1-2 weeks; target K+ 3.5-5.0 mEq/L"],
        ["Simvastatin", "Clarithromycin", "Major", "CYP3A4 inhibition increases Simvastatin levels — risk of myopathy/rhabdomyolysis", "Switch to Azithromycin or hold Simvastatin during antibiotic course"],
        ["Amlodipine", "Simvastatin", "Moderate", "Amlodipine inhibits CYP3A4 — Simvastatin levels may increase", "Limit Simvastatin dose to 20mg/day when combined with Amlodipine"],
        ["Ciprofloxacin", "Antacids (Aluminium/Magnesium)", "Moderate", "Chelation reduces Ciprofloxacin absorption by up to 90%", "Take Ciprofloxacin 2 hours before or 6 hours after antacids"],
        ["Clopidogrel", "Omeprazole", "Moderate", "Omeprazole inhibits CYP2C19 — reduces Clopidogrel antiplatelet effect", "Switch PPI to Pantoprazole (lower CYP2C19 inhibition)"],
        ["Methotrexate", "NSAIDs", "Major", "NSAIDs reduce renal clearance of Methotrexate — toxicity risk", "Avoid NSAIDs; use Paracetamol for analgesia in patients on Methotrexate"],
        ["Lithium", "Thiazide diuretics", "Major", "Diuretics reduce renal Lithium excretion — risk of lithium toxicity", "Monitor Lithium levels weekly; consider alternative diuretic or reduce Lithium dose"],
        ["Sildenafil", "Nitrates", "Contraindicated", "Severe hypotension — both vasodilate via different pathways", "Absolute contraindication; do not co-prescribe under any circumstance"],
        ["Digoxin", "Amiodarone", "Major", "Amiodarone inhibits P-glycoprotein — Digoxin levels increase by 50-100%", "Reduce Digoxin dose by 50% when starting Amiodarone; monitor levels and ECG"],
        ["Tramadol", "SSRIs (e.g., Sertraline)", "Major", "Serotonin syndrome — additive serotonergic effects", "Avoid combination; use alternative analgesics (e.g., low-dose opioids with caution)"],
        ["Fluconazole", "Warfarin", "Major", "CYP2C9 inhibition dramatically increases Warfarin effect", "Reduce Warfarin dose by 30-50%; monitor INR every 2-3 days during fluconazole course"],
        ["Atorvastatin", "Rifampicin", "Major", "Rifampicin induces CYP3A4 — Atorvastatin levels reduced by 80%", "Increase Atorvastatin dose or switch to Rosuvastatin (less CYP3A4-dependent)"],
        ["Phenytoin", "Carbamazepine", "Moderate", "Carbamazepine induces CYP3A4 — may reduce Phenytoin levels", "Monitor Phenytoin levels monthly; adjust dose as required"],
        ["Doxycycline", "Calcium supplements", "Moderate", "Chelation reduces Doxycycline absorption", "Take Doxycycline 1-2 hours before calcium supplements"],
        ["Enoxaparin", "Ketorolac", "Major", "Additive anticoagulant effect — high bleeding risk", "Avoid concurrent use; if pain management needed, use Paracetamol"],
        ["Tacrolimus", "Grapefruit juice", "Moderate", "CYP3A4 inhibition — increases Tacrolimus blood levels unpredictably", "Advise patients to avoid grapefruit and grapefruit juice entirely"],
        ["Metoprolol", "Verapamil", "Major", "Additive negative chronotropic/inotropic effect — bradycardia and heart block risk", "Avoid combination; if essential, monitor ECG continuously and use lowest doses"],
    ]

    out = os.path.join(OUTPUT_DIR, "sample_drug_interactions.csv")
    with open(out, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    print(f"  ✓ CSV created: {out}")


# ─────────────────────────────────────────────────────────────────────────────
# 3. Excel — Patient Records & Prescriptions
# ─────────────────────────────────────────────────────────────────────────────
def make_excel():
    wb = openpyxl.Workbook()

    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
    ALT_FILL     = PatternFill("solid", fgColor="D6E4F0")
    CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def style_sheet(ws, headers):
        ws.row_dimensions[1].height = 28
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font   = HEADER_FONT
            cell.fill   = HEADER_FILL
            cell.alignment = CENTER
        return ws

    # ── Sheet 1: Patient Registry ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Patient Registry"
    h1 = ["Patient ID", "Full Name", "DOB", "Age", "Gender", "Blood Group",
          "Primary Diagnosis", "Secondary Diagnosis", "Allergies", "Attending Physician", "Ward"]
    style_sheet(ws1, h1)
    patients = [
        ["P-1001", "James Mitchell",   "1952-03-14", 72, "M", "A+",  "Hypertension Stage 2", "Type 2 Diabetes",      "Penicillin",        "Dr. Sarah Chen",   "Cardiology"],
        ["P-1002", "Mary Patel",       "1968-07-22", 56, "F", "B+",  "Type 2 Diabetes",      "CKD Stage 3",          "Sulphonamides",     "Dr. Raj Kumar",    "Endocrinology"],
        ["P-1003", "Robert Nguyen",    "1945-11-05", 79, "M", "O-",  "COPD — Gold Stage 3",  "Atrial Fibrillation",  "Aspirin, NSAIDs",   "Dr. Lisa Tran",    "Respiratory"],
        ["P-1004", "Angela Brooks",    "1980-01-30", 45, "F", "AB+", "Rheumatoid Arthritis", "Osteoporosis",         "None known",        "Dr. Tom Walsh",    "Rheumatology"],
        ["P-1005", "David Park",       "1975-09-18", 49, "M", "A-",  "Sepsis — Day 3",       "Pneumonia (CAP)",      "Cephalosporins",    "Dr. Sarah Chen",   "ICU"],
        ["P-1006", "Helen Costa",      "1960-06-03", 64, "F", "B-",  "Heart Failure (EF 35%)","Hypertension",        "ACE inhibitors",    "Dr. Raj Kumar",    "Cardiology"],
        ["P-1007", "Michael Okafor",   "1992-12-27", 32, "M", "O+",  "Epilepsy",             "Depression",           "Carbamazepine",     "Dr. Emma Singh",   "Neurology"],
        ["P-1008", "Susan Lee",        "1958-04-16", 66, "F", "A+",  "Breast Cancer — Stage 2","Anaemia",            "Methotrexate",      "Dr. James Roy",    "Oncology"],
        ["P-1009", "Carlos Reyes",     "1988-08-09", 36, "M", "B+",  "HIV/AIDS — on ART",    "Hepatitis B",          "Doxycycline",       "Dr. Emma Singh",   "Infectious Disease"],
        ["P-1010", "Margaret Johnson", "1940-02-20", 85, "F", "O+",  "Dementia — Alzheimer's","Hypertension, AF",   "Warfarin",          "Dr. Lisa Tran",    "Geriatrics"],
    ]
    for i, row in enumerate(patients, 2):
        for j, val in enumerate(row, 1):
            c = ws1.cell(row=i, column=j, value=val)
            if i % 2 == 0:
                c.fill = ALT_FILL
    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["G"].width = 28
    ws1.column_dimensions["H"].width = 22
    ws1.column_dimensions["I"].width = 22

    # ── Sheet 2: Current Prescriptions ────────────────────────────────────
    ws2 = wb.create_sheet("Prescriptions")
    h2 = ["Patient ID", "Medication", "Dose", "Route", "Frequency", "Indication",
          "Start Date", "Review Date", "Prescribing Doctor", "Notes"]
    style_sheet(ws2, h2)
    prescriptions = [
        ["P-1001", "Amlodipine",         "10mg",    "Oral", "Once daily",       "Stage 2 HTN",        "2025-01-10", "2025-04-10", "Dr. Sarah Chen", "Monitor ankle oedema"],
        ["P-1001", "Metformin",          "1000mg",  "Oral", "Twice daily",      "Type 2 DM",          "2025-01-10", "2025-04-10", "Dr. Sarah Chen", "Take with meals"],
        ["P-1002", "Insulin Glargine",   "20 units","SC",   "At bedtime",       "Type 2 DM — poor control","2025-02-01","2025-05-01","Dr. Raj Kumar","Adjust dose per SMBG"],
        ["P-1002", "Losartan",           "50mg",    "Oral", "Once daily",       "HTN + CKD renoprotection","2025-02-01","2025-05-01","Dr. Raj Kumar","Monitor K+ monthly"],
        ["P-1003", "Tiotropium inhaler", "18mcg",   "Inhaled","Once daily",     "COPD — bronchodilation","2024-11-15","2025-05-15","Dr. Lisa Tran","Rinse mouth after use"],
        ["P-1003", "Edoxaban",           "30mg",    "Oral", "Once daily",       "Atrial Fibrillation","2024-11-15","2025-05-15","Dr. Lisa Tran","Reduce dose for low weight/CrCl"],
        ["P-1004", "Methotrexate",       "15mg",    "Oral", "Weekly (Monday)",  "Rheumatoid Arthritis","2024-09-01","2025-03-01","Dr. Tom Walsh","MUST take Folic Acid 5mg 6 days/week"],
        ["P-1004", "Folic Acid",         "5mg",     "Oral", "Six days/week",    "MTX side-effect prevention","2024-09-01","2025-03-01","Dr. Tom Walsh","Take every day EXCEPT MTX day"],
        ["P-1005", "Meropenem",          "1g",      "IV",   "Every 8 hours",    "Sepsis / CAP",       "2025-03-01", "2025-03-08", "Dr. Sarah Chen", "Day 3 of 7-day course; culture-guided de-escalation"],
        ["P-1005", "Norepinephrine",     "0.1 mcg/kg/min","IV drip","Continuous","Septic shock — MAP support","2025-03-01","2025-03-03","Dr. Sarah Chen","Titrate to MAP >= 65 mmHg"],
        ["P-1006", "Furosemide",         "40mg",    "Oral", "Twice daily",      "Heart Failure — fluid overload","2025-01-20","2025-04-20","Dr. Raj Kumar","Monitor daily weights; target -0.5 to -1 kg/day"],
        ["P-1006", "Carvedilol",         "6.25mg",  "Oral", "Twice daily",      "HF — EF reduction",  "2025-01-20", "2025-04-20", "Dr. Raj Kumar", "Uptitrate slowly every 2 weeks"],
        ["P-1007", "Levetiracetam",      "1000mg",  "Oral", "Twice daily",      "Epilepsy control",   "2024-06-15", "2025-06-15", "Dr. Emma Singh","Review seizure diary at each visit"],
        ["P-1008", "Paclitaxel",         "175 mg/m2","IV",  "Every 3 weeks",    "Breast Cancer — adjuvant chemo","2025-02-10","2025-05-10","Dr. James Roy","Pre-medicate: Dexamethasone + Diphenhydramine"],
        ["P-1009", "Tenofovir/Emtricitabine","300/200mg","Oral","Once daily",   "HIV ART",            "2023-07-01", "2025-07-01", "Dr. Emma Singh","Renal function every 6 months"],
        ["P-1010", "Donepezil",          "10mg",    "Oral", "At bedtime",       "Alzheimer's dementia","2024-03-01","2025-03-01","Dr. Lisa Tran", "Watch for bradycardia — check pulse monthly"],
        ["P-1010", "Apixaban",           "2.5mg",   "Oral", "Twice daily",      "Atrial Fibrillation","2024-03-01","2025-03-01","Dr. Lisa Tran", "Reduced dose: age >80, low weight"],
    ]
    for i, row in enumerate(prescriptions, 2):
        for j, val in enumerate(row, 1):
            c = ws2.cell(row=i, column=j, value=val)
            if i % 2 == 0:
                c.fill = ALT_FILL
    for col in ["B","F","J"]:
        ws2.column_dimensions[col].width = 28

    # ── Sheet 3: Lab Results ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Lab Results")
    h3 = ["Patient ID", "Test Name", "Result", "Unit", "Reference Range", "Flag", "Date", "Ordering Doctor"]
    style_sheet(ws3, h3)
    labs = [
        ["P-1001", "HbA1c",         "8.2",  "%",          "4.0–5.6",          "HIGH", "2025-02-28", "Dr. Sarah Chen"],
        ["P-1001", "Serum Creatinine","112", "umol/L",     "62–115",           "Normal","2025-02-28","Dr. Sarah Chen"],
        ["P-1001", "eGFR",          "58",   "mL/min/1.73m2",">=60",           "LOW",  "2025-02-28", "Dr. Sarah Chen"],
        ["P-1002", "HbA1c",         "10.1", "%",          "4.0–5.6",          "HIGH", "2025-02-20", "Dr. Raj Kumar"],
        ["P-1002", "Serum Potassium","5.6", "mEq/L",      "3.5–5.0",          "HIGH", "2025-02-20", "Dr. Raj Kumar"],
        ["P-1002", "eGFR",          "22",   "mL/min/1.73m2",">=60",           "CRITICAL LOW","2025-02-20","Dr. Raj Kumar"],
        ["P-1003", "FEV1",          "42",   "% predicted",">=80",             "LOW",  "2025-01-15", "Dr. Lisa Tran"],
        ["P-1003", "INR",           "2.4",  "ratio",      "2.0–3.0 (AF)",     "Normal","2025-02-25","Dr. Lisa Tran"],
        ["P-1005", "Lactate",       "4.8",  "mmol/L",     "<2.0",             "CRITICAL HIGH","2025-03-01","Dr. Sarah Chen"],
        ["P-1005", "WBC",           "22.4", "x10^9/L",    "4.0–11.0",         "HIGH", "2025-03-01", "Dr. Sarah Chen"],
        ["P-1005", "CRP",           "286",  "mg/L",       "<5",               "HIGH", "2025-03-01", "Dr. Sarah Chen"],
        ["P-1006", "BNP",           "1820", "pg/mL",      "<100",             "CRITICAL HIGH","2025-01-18","Dr. Raj Kumar"],
        ["P-1006", "Serum Sodium",  "128",  "mEq/L",      "135–145",          "LOW",  "2025-01-18", "Dr. Raj Kumar"],
        ["P-1008", "Neutrophil count","0.8","x10^9/L",   ">2.0",             "CRITICAL LOW","2025-02-25","Dr. James Roy"],
        ["P-1010", "Serum Digoxin", "2.8",  "nmol/L",     "1.0–2.6",         "HIGH", "2025-02-10", "Dr. Lisa Tran"],
    ]
    for i, row in enumerate(labs, 2):
        for j, val in enumerate(row, 1):
            c = ws3.cell(row=i, column=j, value=val)
            if row[5] in ("CRITICAL HIGH", "CRITICAL LOW"):
                c.fill = PatternFill("solid", fgColor="FFD7D7")
            elif i % 2 == 0:
                c.fill = ALT_FILL
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["E"].width = 22

    out = os.path.join(OUTPUT_DIR, "sample_patient_records.xlsx")
    wb.save(out)
    print(f"  ✓ Excel created: {out}")


# ─────────────────────────────────────────────────────────────────────────────
# 4. TXT — Doctor's Clinical Notes
# ─────────────────────────────────────────────────────────────────────────────
def make_txt():
    content = """\
================================================================================
CITY GENERAL HOSPITAL — CLINICAL NOTES REPOSITORY
Department of Internal Medicine & Critical Care
Generated: March 2025
================================================================================

────────────────────────────────────────────────────────────────────────────────
PATIENT: James Mitchell | ID: P-1001 | DOB: 14-Mar-1952 | Ward: Cardiology
ATTENDING: Dr. Sarah Chen | Date: 28-Feb-2025
────────────────────────────────────────────────────────────────────────────────

PRESENTING COMPLAINT:
Patient presents for routine follow-up of Stage 2 hypertension and Type 2 Diabetes
Mellitus. Reports intermittent headaches in the morning, rated 4/10. Denies chest pain,
palpitations, or visual disturbances. Good medication compliance per patient report.

VITAL SIGNS:
  BP: 158/96 mmHg (elevated — target <130/80)  HR: 78 bpm  SpO2: 97%  Temp: 36.8°C
  Weight: 94 kg  BMI: 31.2 (Obese Class I)

ASSESSMENT & PLAN:
  1. Hypertension — suboptimally controlled. Increase Amlodipine from 5mg to 10mg OD.
     Add Indapamide 1.5mg MR if BP remains >140/90 at next visit in 4 weeks.
  2. Type 2 DM — HbA1c 8.2% (target <7%). Increase Metformin to 1000mg BD.
     Refer to dietitian for medical nutrition therapy. Consider adding Empagliflozin
     10mg OD at next visit given BMI and cardiovascular risk.
  3. CKD Stage 2 (eGFR 58) — monitor. AVOID NSAIDs. Annual urine ACR.
  4. Statin therapy — Atorvastatin 40mg OD continued. LDL target <1.8 mmol/L.
  REVIEW: 4 weeks. FBC, U&E, HbA1c, fasting lipids to be repeated.

────────────────────────────────────────────────────────────────────────────────
PATIENT: David Park | ID: P-1005 | DOB: 18-Sep-1975 | Ward: ICU (Bed 4)
ATTENDING: Dr. Sarah Chen | Date: 01-Mar-2025 | Time: 06:30
────────────────────────────────────────────────────────────────────────────────

PRESENTING COMPLAINT:
49-year-old male admitted via Emergency Dept with 3-day history of productive cough,
fever (Tmax 39.4°C), and rigors. Developed sudden haemodynamic deterioration at 02:00.
BP dropped to 72/40 mmHg. Diagnosed as Septic Shock secondary to Community-Acquired
Pneumonia. History of allergy to Cephalosporins (rash).

MANAGEMENT (Hour-1 Sepsis Bundle — COMPLETED at 03:15):
  ✓ Blood cultures x2 taken at 02:45 — results pending
  ✓ Lactate: 4.8 mmol/L (critical — repeat at 06:00)
  ✓ IV access: 2x large-bore venflon + Central venous catheter (right subclavian)
  ✓ IV Meropenem 1g q8h commenced 03:00 (Day 1)
  ✓ IV Norepinephrine 0.1 mcg/kg/min commenced — MAP now 66 mmHg
  ✓ 30 mL/kg crystalloid bolus (2L 0.9% NaCl) over 30 min — completed
  ✓ Chest X-Ray: Right lower lobe consolidation consistent with pneumonia
  ✓ ABG: pH 7.28, pCO2 33, pO2 68, HCO3 15 — metabolic acidosis with respiratory compensation

CURRENT STATUS (06:30 round):
  MAP: 67 mmHg on Norepinephrine 0.12 mcg/kg/min
  SpO2: 91% on High Flow Nasal Cannula (HFNC) 60L/min, FiO2 0.5
  Urine output: 35 mL/hour (adequate)
  Lactate repeat (06:00): 3.2 mmol/L — trending down, reassuring
  
  PLAN: Continue Meropenem. Low threshold for intubation if SpO2 deteriorates below 88%.
  Microbiology to call with sensitivities — de-escalate antibiotics when available.
  Echo at 08:00 to assess cardiac function. ICU consultant review at 09:00.
  Family updated at 07:00 — understand severity; wife consent obtained for ICU level care.

────────────────────────────────────────────────────────────────────────────────
PATIENT: Helen Costa | ID: P-1006 | DOB: 03-Jun-1960 | Ward: Cardiology
ATTENDING: Dr. Raj Kumar | Date: 18-Jan-2025
────────────────────────────────────────────────────────────────────────────────

PRESENTING COMPLAINT:
64F with known ischaemic cardiomyopathy (EF 35%) presents with 5-day history of
progressive dyspnoea (now NYHA Class III), bilateral leg swelling to knees, and
orthopnoea (sleeps on 3 pillows). Weight gain of 5 kg over 10 days.

VITAL SIGNS:
  BP: 102/68 mmHg  HR: 102 bpm (irregular — AF)  SpO2: 89% on room air
  JVP elevated 6 cm. Bilateral crackles to mid-zones. Pitting oedema 2+ to knees.

ECG: Atrial fibrillation with rapid ventricular response, rate 105 bpm. Widespread
     ST depression — consistent with demand ischaemia.

ECHO (today): EF 33% (deteriorated from 38% in October). Moderate mitral regurgitation.
              LV hypokinesis — diffuse. No pericardial effusion.

ASSESSMENT & PLAN:
  Diagnosis: Acute decompensated heart failure with reduced EF (HFrEF)
  
  1. IV Furosemide 80mg twice daily — target diuresis 1.5–2L/day. Daily weights.
     Strict fluid restriction 1.5L/day. Daily U&E — watch for hypokalaemia.
  2. Carvedilol 6.25mg BD — continue but HOLD if SBP falls below 90 mmHg.
  3. AF with RVR — IV Digoxin loading 0.5mg now, then 0.25mg in 6 hours.
     Maintain Apixaban anticoagulation. Target resting HR <80 bpm.
  4. Hyponatraemia (Na 128) — secondary to fluid overload + neurohormonal activation.
     Restrict free water. Avoid Hypotonic IV fluids. Repeat lytes in 8 hours.
  5. Cardiac rehabilitation referral once stabilised.
  REVIEW: Twice daily. Cardiology consultant ward round 08:00.

────────────────────────────────────────────────────────────────────────────────
PATIENT: Margaret Johnson | ID: P-1010 | DOB: 20-Feb-1940 | Ward: Geriatrics
ATTENDING: Dr. Lisa Tran | Date: 10-Feb-2025
────────────────────────────────────────────────────────────────────────────────

PRESENTING COMPLAINT:
85F with Alzheimer's dementia, hypertension, and atrial fibrillation on Apixaban and
Donepezil. Brought in by family; confused and uncooperative. Serum Digoxin level 2.8
nmol/L (above therapeutic range 1.0–2.6). Last Digoxin dose unknown — possible
inadvertent double dosing at nursing home.

CLINICAL ASSESSMENT:
  - Mild Digoxin toxicity: nausea, vomiting, visual disturbance (yellow-green halos)
  - Bradycardia: HR 48 bpm, BP 105/65 mmHg, SpO2 96%
  - ECG: Sinus bradycardia, PR prolongation (240ms). No heart block. No ventricular ectopy.

MANAGEMENT:
  1. HOLD Digoxin — do not give further doses until repeat level in 24h is <2.0 nmol/L.
  2. Continuous cardiac monitoring — alert for AV block or VT.
  3. IV access. Check electrolytes — hypokalaemia worsens Digoxin toxicity.
     K+ 3.1 (low) — IV potassium replacement 40 mmol in 500ml NaCl over 4 hours.
  4. Digoxin-specific antibody fragments (DigiFab) — NOT indicated at this level;
     threshold for DigiFab: life-threatening arrhythmias or K+ >5.5 with toxicity.
  5. Contact nursing home pharmacist — review medication administration records.
     Notify Pharmacy for clinical review of drug chart.
  REVIEW: 4-hourly obs. Repeat ECG in 4 hours. Senior review if HR <40 or AV block.

================================================================================
END OF CLINICAL NOTES — CONFIDENTIAL — FOR AUTHORISED HEALTHCARE PERSONNEL ONLY
================================================================================
"""
    out = os.path.join(OUTPUT_DIR, "sample_treatment_notes.txt")
    with open(out, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"  ✓ TXT created: {out}")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Generating healthcare sample test files...")
    make_pdf()
    make_csv()
    make_excel()
    make_txt()
    print(f"\nAll files saved to: {OUTPUT_DIR}/")
